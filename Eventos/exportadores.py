# exportadores.py
#
# Lógica de generación de reportes, independiente de Django's request/response.
# No importa nada de django.http aquí: solo recibe datos y devuelve un Workbook.
# Esto permite reutilizarla desde una vista, un management command, un task de
# Celery (ej. para enviar el Excel por correo), o probarla con un test unitario
# simple sin necesidad de un RequestFactory.

import logging
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

logger = logging.getLogger(__name__)


def _preparar_logo_excel(organizacion, alto_px=70):
    """
    Lee el logo desde el storage de Django (funciona con FileSystemStorage,
    S3, o cualquier backend), lo redimensiona manteniendo el aspect ratio,
    y devuelve un openpyxl.drawing.image.Image listo para ws.add_image().

    Devuelve None si no hay logo o si no se puede leer (archivo corrupto,
    borrado del storage, etc.) — nunca debe tumbar la generación del reporte.
    """
    if not organizacion or not organizacion.logo:
        return None

    try:
        organizacion.logo.open("rb")
        datos_originales = organizacion.logo.read()
    except Exception:
        logger.warning(
            "No se pudo leer el logo de la organización %s para el reporte Excel",
            getattr(organizacion, "pk", "?"),
            exc_info=True,
        )
        return None
    finally:
        try:
            organizacion.logo.close()
        except Exception:
            pass

    try:
        imagen = PILImage.open(BytesIO(datos_originales))
        imagen = imagen.convert("RGBA")

        ancho_original, alto_original = imagen.size
        ratio = alto_px / float(alto_original)
        ancho_px = int(ancho_original * ratio)
        imagen = imagen.resize((ancho_px, alto_px), PILImage.LANCZOS)

        buffer_redimensionado = BytesIO()
        imagen.save(buffer_redimensionado, format="PNG")
        buffer_redimensionado.seek(0)

        xl_img = XLImage(buffer_redimensionado)
        xl_img.width = ancho_px
        xl_img.height = alto_px
        return xl_img
    except Exception:
        logger.warning(
            "El logo de la organización %s no es una imagen válida",
            getattr(organizacion, "pk", "?"),
            exc_info=True,
        )
        return None


def generar_excel_asistencia(actividad, asistencias, organizacion=None, usuario=None):
    """
    Construye el Workbook del reporte de asistencia.

    Parámetros
    ----------
    actividad : Actividad
        Instancia del evento/actividad.
    asistencias : QuerySet[RegistroAsistencia]
        Queryset YA optimizado (select_related/prefetch_related) y filtrado
        por la vista que llama a esta función.
    organizacion : Organizacion | None
        Para el encabezado institucional. Opcional.
    usuario : User | None
        Quién generó el reporte, para el pie/encabezado. Opcional.

    Retorna
    -------
    openpyxl.Workbook
        Listo para wb.save(...) sobre un archivo o un HttpResponse.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # --- Estilos reutilizables -----------------------------------------
    FONT_TITULO = Font(name="Arial", size=14, bold=True, color="1A365D")
    FONT_SUBTITULO = Font(name="Arial", size=9, color="666666")
    FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    FILL_HEADER = PatternFill("solid", fgColor="1A365D")
    FONT_CONFIRMADO = Font(name="Arial", size=10, bold=True, color="0F5132")
    FILL_CONFIRMADO = PatternFill("solid", fgColor="D1E7DD")
    FONT_REGISTRADO = Font(name="Arial", size=10, color="41464B")
    FILL_REGISTRADO = PatternFill("solid", fgColor="E2E3E5")
    BORDE_FINO = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
    ALIGN_WRAP = Alignment(vertical="center", wrap_text=True)

    # --- Encabezado institucional (con logo, si existe) -------------------
    nombre_org = organizacion.nombre_organizacion if organizacion else "Institución del Sistema"

    logo = _preparar_logo_excel(organizacion)

    if logo:
        # Columna A reservada para el logo → el título arranca en B y se
        # extiende hasta I para cubrir las 9 columnas de datos.
        ws.add_image(logo, "A1")
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 16
        ws.column_dimensions["A"].width = max(
            ws.column_dimensions["A"].width or 0, (logo.width / 7) + 2
        )
        col_inicio, col_fin = "B", "I"
    else:
        col_inicio, col_fin = "A", "G"

    ws.merge_cells(f"{col_inicio}1:{col_fin}1")
    ws[f"{col_inicio}1"] = nombre_org
    ws[f"{col_inicio}1"].font = FONT_TITULO
    ws[f"{col_inicio}1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"{col_inicio}2:{col_fin}2")
    ws[f"{col_inicio}2"] = f"Reporte de Asistencia — {actividad.nombre_evento}"
    ws[f"{col_inicio}2"].font = FONT_SUBTITULO
    ws[f"{col_inicio}2"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"{col_inicio}3:{col_fin}3")
    generado_por = f" | Generado por: {usuario}" if usuario else ""
    ws[f"{col_inicio}3"] = f"Fecha de impresión: {datetime.now().strftime('%d/%m/%Y %H:%M')}{generado_por}"
    ws[f"{col_inicio}3"].font = FONT_SUBTITULO
    ws[f"{col_inicio}3"].alignment = Alignment(horizontal="center")

    ws.append([])  # separador

    # --- Encabezados de columnas ----------------------------------------
    columnas = [
        "Identificación",
        "Nombre Completo",
        "Correo Electrónico",
        "Organización de Origen",
        "Seudónimo",
        "Cargos Asignados",
        "Estado",
        "Fecha de Registro",
        "Fecha de Confirmación",
    ]
    ws.append(columnas)
    fila_header = ws.max_row
    for col_idx in range(1, len(columnas) + 1):
        celda = ws.cell(row=fila_header, column=col_idx)
        celda.font = FONT_HEADER
        celda.fill = FILL_HEADER
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = BORDE_FINO

    # --- Filas de datos ---------------------------------------------------
    for reg in asistencias:
        cargos = ", ".join(
            pc.cargo.nombre_cargo for pc in reg.asistente.personacargo_set.all()
        ) or "Sin cargos asignados"

        fila = [
            reg.asistente.identificacion,
            f"{reg.asistente.nombres} {reg.asistente.apellidos}",
            reg.asistente.email,
            reg.organizacion_origen or "Particular/Externo",
            reg.seudonimo or "",
            cargos,
            "Confirmado" if reg.estado == "CONFIRMADO" else "Registrado",
            reg.fecha_registro.strftime("%d/%m/%Y %H:%M") if reg.fecha_registro else "",
            reg.fecha_confirmacion.strftime("%d/%m/%Y %H:%M")
            if getattr(reg, "fecha_confirmacion", None) else "",
        ]
        ws.append(fila)

        fila_actual = ws.max_row
        es_confirmado = reg.estado == "CONFIRMADO"
        fuente = FONT_CONFIRMADO if es_confirmado else FONT_REGISTRADO
        relleno = FILL_CONFIRMADO if es_confirmado else FILL_REGISTRADO

        for col_idx in range(1, len(columnas) + 1):
            celda = ws.cell(row=fila_actual, column=col_idx)
            celda.border = BORDE_FINO
            celda.alignment = ALIGN_WRAP
            if col_idx == 7:
                celda.font = fuente
                celda.fill = relleno

    # --- Presentación -------------------------------------------------
    anchos = [16, 28, 30, 24, 18, 26, 14, 18, 18]
    for i, ancho in enumerate(anchos, start=1):
        letra = get_column_letter(i)
        # No reducimos la columna A por debajo de lo que necesita el logo
        ancho_actual = ws.column_dimensions[letra].width
        ws.column_dimensions[letra].width = max(ancho, ancho_actual or 0)

    ws.freeze_panes = f"A{fila_header + 1}"
    ws.auto_filter.ref = f"A{fila_header}:{get_column_letter(len(columnas))}{ws.max_row}"

    return wb